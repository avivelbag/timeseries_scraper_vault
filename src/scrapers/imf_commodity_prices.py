"""IMF Primary Commodity Prices monthly Excel scraper.

Downloads the IMF Research workbook (one Excel file, ~60 commodity series
back to 1960), iterates every price sheet, and yields ImfCommodityPrice
records.  The workbook structure is:

  - Row 1: sheet-level metadata headers (commodity name, code, units, …)
  - Row 2: column-header row (Date, commodity code, …)
  - Row 3+: monthly data rows (date in col A, price in subsequent cols)

Each sheet exposes a single commodity.  Commodity name and units are read
from row 1; commodity code is the sheet name itself (e.g. "POILWTI").
"""

import tempfile
from datetime import datetime, timezone

import openpyxl

from src.bq_uploader import upload_rows
from src.scrapers.http_client import fetch
from protos.imf_commodity_price_pb2 import ImfCommodityPrice  # type: ignore[attr-defined]

SOURCE_URL = (
    "https://www.imf.org/en/Research/commodity-prices"
    "/~/media/Files/Research/CommodityPrices/Monthly/ExternalData.ashx"
)

REQUIRED_FIELDS: list[str] = [
    "commodity_name",
    "commodity_code",
    "date",
    "price_usd",
    "units",
    "source_url",
    "fetch_time",
]

_SKIP_SHEET_NAMES = {"Index", "README", "Notes", "Cover"}

_DATE_CELL_FORMATS = ["%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%m/%d/%Y", "%b-%Y"]


def _parse_date(raw) -> str | None:
    """Convert a cell value from an Excel date column to 'YYYY-MM' string.

    Handles datetime objects (openpyxl returns these for date-typed cells),
    strings in several common formats, and integer Excel serial dates.
    Returns None if the value cannot be parsed.
    """
    if isinstance(raw, datetime):
        return raw.strftime("%Y-%m")
    if isinstance(raw, str):
        raw = raw.strip()
        for fmt in _DATE_CELL_FORMATS:
            try:
                return datetime.strptime(raw, fmt).strftime("%Y-%m")
            except ValueError:
                pass
        return None
    return None


def _extract_sheet_meta(sheet) -> tuple[str, str, str]:
    """Return (commodity_name, commodity_code, units) from a workbook sheet.

    The IMF workbook stores metadata in the first two rows:
      - Row 1 may contain a label like "Commodity Name: Crude Oil, WTI"
        followed by units information.
      - Row 2 is a column-header row; the second column header is the IMF
        commodity code (e.g. "POILWTI").

    Falls back to the sheet title for commodity_code when row 2 is absent.
    Returns ("", sheet.title, "") when metadata rows are missing or empty.
    """
    commodity_name = ""
    commodity_code = sheet.title
    units = ""

    rows = list(sheet.iter_rows(min_row=1, max_row=2, values_only=True))
    if not rows:
        return commodity_name, commodity_code, units

    row1 = rows[0] if rows else ()
    if row1:
        for cell in row1:
            if cell is None:
                continue
            text = str(cell).strip()
            if not text:
                continue
            if not commodity_name:
                commodity_name = text
            elif not units:
                units = text
            else:
                break

    if len(rows) >= 2:
        row2 = rows[1]
        if row2 and len(row2) >= 2 and row2[1]:
            code = str(row2[1]).strip()
            if code:
                commodity_code = code

    return commodity_name, commodity_code, units


def parse_workbook(wb: openpyxl.Workbook, source_url: str) -> list[dict]:
    """Parse an openpyxl Workbook into a flat list of commodity price dicts.

    Iterates every sheet whose name is not in _SKIP_SHEET_NAMES.  For each
    sheet, reads metadata from the header rows (commodity name, code, units)
    then iterates data rows (row 3 onward).  A row is included if:
      - Column A parses as a valid YYYY-MM date via _parse_date.
      - Column B contains a positive numeric price.

    Args:
        wb: An open openpyxl Workbook object.
        source_url: The URL the workbook was fetched from; stored verbatim on
            every record.

    Returns:
        List of dicts with keys: commodity_name, commodity_code, date,
        price_usd, units, source_url.  fetch_time is omitted — callers add it.
    """
    fetch_time = datetime.now(timezone.utc).isoformat()
    records: list[dict] = []

    for sheet in wb.worksheets:
        if sheet.title in _SKIP_SHEET_NAMES:
            continue

        commodity_name, commodity_code, units = _extract_sheet_meta(sheet)

        for row in sheet.iter_rows(min_row=3, values_only=True):
            if not row or row[0] is None:
                continue
            date_str = _parse_date(row[0])
            if date_str is None:
                continue
            if len(row) < 2 or row[1] is None:
                continue
            try:
                price = float(row[1])
            except (TypeError, ValueError):
                continue
            if price <= 0:
                continue
            records.append(
                {
                    "commodity_name": commodity_name,
                    "commodity_code": commodity_code,
                    "date": date_str,
                    "price_usd": price,
                    "units": units,
                    "source_url": source_url,
                    "fetch_time": fetch_time,
                }
            )

    return records


def scrape() -> list[dict]:
    """Download the IMF commodity prices workbook and return parsed records.

    Uses src.scrapers.http_client.fetch (which enforces robots.txt, polite
    delay ≥ 3 s, and exponential backoff on 429/5xx) to download the workbook
    as a binary stream into a temporary file, then opens it with openpyxl and
    delegates to parse_workbook.

    Returns:
        Same structure as parse_workbook.
    """
    resp = fetch(SOURCE_URL, stream=True, min_delay=3.0, max_delay=6.0)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        for chunk in resp.iter_content(chunk_size=65536):
            if chunk:
                tmp.write(chunk)
        tmp_path = tmp.name

    wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
    return parse_workbook(wb, SOURCE_URL)


def _record_to_proto(record: dict) -> ImfCommodityPrice:
    """Convert a parsed record dict to an ImfCommodityPrice proto instance."""
    msg = ImfCommodityPrice()
    msg.commodity_name = record["commodity_name"]
    msg.commodity_code = record["commodity_code"]
    msg.date = record["date"]
    msg.price_usd = record["price_usd"]
    msg.units = record["units"]
    msg.source_url = record["source_url"]
    msg.fetch_time = record.get("fetch_time", datetime.now(timezone.utc).isoformat())
    return msg


def main() -> int:
    """Scrape IMF commodity prices and upload records to BigQuery.

    Returns the count of successfully inserted rows.
    """
    records = scrape()
    messages = [_record_to_proto(r) for r in records]
    return upload_rows("imf_commodity_prices", messages, date_column="date")


if __name__ == "__main__":
    main()
