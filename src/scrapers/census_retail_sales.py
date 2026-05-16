"""Census Bureau Advance Monthly Retail Trade and Food Services scraper.

Fetches the Census advance retail Excel workbook (marts_current.xlsx), which is
the documented static endpoint linked from the timeseries and advance-release
pages. The HTML at timeseries.html is JavaScript-rendered and carries no static
table; the Excel is the authoritative machine-readable source.

Table 1 holds seasonally adjusted sales (millions USD) with columns labelled
(a) advance, (p) preliminary, and (r) revised. Table 2 holds the corresponding
month-over-month and year-over-year percent changes for the advance and
preliminary periods. One record is emitted per (series, period) pair.
"""

import io
import re
import time
from datetime import datetime, timezone

import openpyxl

from protos.census_retail_sales_pb2 import CensusRetailSalesRecord  # type: ignore[attr-defined]
from src.bq_uploader import upload_rows
from src.scrapers.http_client import fetch

SOURCE_URL = "https://www.census.gov/retail/marts/www/marts_current.xlsx"

_MONTH_ABBR: dict[str, int] = {
    "Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4, "May": 5, "Jun": 6,
    "Jul": 7, "Aug": 8, "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12,
}

# Known column layout in the Census advance retail Excel workbook (0-based).
# Rows 6-8 (Excel rows 7-9) are the three-row header band.
_YEAR_ROW_IDX = 6
_MONTH_ROW_IDX = 7
_MARKER_ROW_IDX = 8
_NAICS_COL = 0
_NAME_COL = 1

# Seasonally adjusted data columns in Table 1 (Excel columns J-N, 0-based 9-13).
_ADJ_COLS = [9, 10, 11, 12, 13]

# Percent-change columns in Table 2 (0-based):
# col 2 = advance MoM, col 3 = advance YoY, col 4 = prelim MoM, col 5 = prelim YoY
_T2_ADV_MOM = 2
_T2_ADV_YOY = 3
_T2_PRE_MOM = 4
_T2_PRE_YOY = 5

_DOT_LEADER = re.compile(r"[.…\s]+$")

# The Census Excel has a fixed header band (rows 1-10, 0-based 0-9).
# Row 10 (idx 9) is a blank separator; data starts at row 11 (idx 10).
_DATA_START_ROW = 10


def _clean_name(raw: object) -> str:
    """Strip Census dot-leaders and surrounding whitespace from a series label."""
    if not raw:
        return ""
    return _DOT_LEADER.sub("", str(raw)).strip()


def _parse_numeric(val: object) -> float | None:
    """Return float for a numeric cell, None for suppressed or missing values.

    Census marks suppressed data as '(*)', not-applicable as '(NA)', and
    withheld-to-avoid-disclosure as '(S)'. Any of these map to None.
    """
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if s in ("", "-", "--", "(NA)", "(*)", "(S)", " "):
        return None
    s = s.replace(",", "")
    try:
        return float(s)
    except ValueError:
        return None


def _parse_col_meta(rows: list[tuple], col_indices: list[int]) -> list[dict]:
    """Extract period_date and revised status for each column index.

    Reads the three-row header band (year row, month row, marker row) from the
    worksheet. The year value carries forward from left to right within a
    section, so None in a year cell inherits the previous column's year.

    Args:
        rows: All rows from the worksheet as tuples of cell values.
        col_indices: 0-based column indices to inspect.

    Returns:
        List of dicts with keys 'col_idx', 'period_date' (YYYY-MM-01),
        and 'revised' (bool). Columns that cannot be parsed are omitted.
    """
    year_row = rows[_YEAR_ROW_IDX] if len(rows) > _YEAR_ROW_IDX else ()
    month_row = rows[_MONTH_ROW_IDX] if len(rows) > _MONTH_ROW_IDX else ()
    marker_row = rows[_MARKER_ROW_IDX] if len(rows) > _MARKER_ROW_IDX else ()

    last_year: int | None = None
    result: list[dict] = []

    for col_idx in sorted(col_indices):
        year_val = year_row[col_idx] if col_idx < len(year_row) else None
        if isinstance(year_val, int) and year_val > 1900:
            last_year = year_val
        year = last_year

        month_raw = month_row[col_idx] if col_idx < len(month_row) else None
        month_str = re.sub(r"[^A-Za-z]", "", str(month_raw or ""))[:3]
        month_num = _MONTH_ABBR.get(month_str.capitalize())

        marker_val = marker_row[col_idx] if col_idx < len(marker_row) else None
        revised = str(marker_val or "").strip() == "(r)"

        if year and month_num:
            result.append({
                "col_idx": col_idx,
                "period_date": f"{year:04d}-{month_num:02d}-01",
                "revised": revised,
            })

    return result


def _load_table(wb: openpyxl.Workbook, sheet_name: str) -> list[tuple]:
    """Load all rows from a named worksheet as a list of value tuples."""
    ws = wb[sheet_name]
    return [tuple(cell.value for cell in row) for row in ws.iter_rows()]


def _index_pct_table(rows: list[tuple]) -> dict[str, dict]:
    """Build a lookup from full series name to percent-change values from Table 2.

    Applies the same multi-line name concatenation used in run(): when a row
    has a name but no numeric pct values, its name becomes the prefix for the
    next data row. Returns a dict keyed by the fully assembled series name.

    Args:
        rows: All rows from Table 2.

    Returns:
        Dict mapping cleaned series_name →
            {'adv_mom': float|None, 'adv_yoy': float|None,
             'pre_mom': float|None, 'pre_yoy': float|None}.
    """
    index: dict[str, dict] = {}
    pending_prefix = ""
    for row in rows:
        if len(row) <= _T2_PRE_YOY:
            pending_prefix = ""
            continue

        name_raw = row[_NAME_COL]
        adv_mom = _parse_numeric(row[_T2_ADV_MOM])
        adv_yoy = _parse_numeric(row[_T2_ADV_YOY])
        pre_mom = _parse_numeric(row[_T2_PRE_MOM])
        pre_yoy = _parse_numeric(row[_T2_PRE_YOY])
        has_pct = any(v is not None for v in (adv_mom, adv_yoy, pre_mom, pre_yoy))

        if name_raw and not has_pct:
            pending_prefix = _clean_name(str(name_raw))
            continue

        if not name_raw and not has_pct:
            pending_prefix = ""
            continue

        name_part = _clean_name(str(name_raw)) if name_raw else ""
        if pending_prefix and name_part:
            name = pending_prefix + " " + name_part
        elif pending_prefix:
            name = pending_prefix
        else:
            name = name_part
        pending_prefix = ""

        name = _DOT_LEADER.sub("", name).strip()
        if not name:
            continue

        index[name] = {
            "adv_mom": adv_mom,
            "adv_yoy": adv_yoy,
            "pre_mom": pre_mom,
            "pre_yoy": pre_yoy,
        }
    return index


def run(xlsx_bytes: bytes, source_url: str = SOURCE_URL) -> list[dict]:
    """Parse a Census advance retail Excel workbook into per-(series, period) records.

    Reads Table 1 for seasonally adjusted sales figures and Table 2 for
    month-over-month and year-over-year percent changes. Emits one record per
    (NAICS series, period) pair where the adjusted sales value is present.

    The revised flag is set to True for columns labelled (r) in the workbook
    header band. Percent changes are included for advance and preliminary
    periods only (Table 2 does not publish changes for the revised column).

    Args:
        xlsx_bytes: Raw bytes of the marts_current.xlsx file.
        source_url: URL the workbook was fetched from, stored in each record.

    Returns:
        List of dicts with keys: series_name, period_date, sales_millions_usd,
        month_over_month_pct, year_over_year_pct, revised, source_url.

    Raises:
        ValueError: When xlsx_bytes is empty or contains no parseable records.
    """
    if not xlsx_bytes:
        raise ValueError("Empty Excel bytes provided to Census retail sales parser")

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))

    if "Table 1." not in wb.sheetnames or "Table 2." not in wb.sheetnames:
        raise ValueError(
            f"Expected sheets 'Table 1.' and 'Table 2.' not found; got: {wb.sheetnames}"
        )

    t1_rows = _load_table(wb, "Table 1.")
    t2_rows = _load_table(wb, "Table 2.")

    col_meta = _parse_col_meta(t1_rows, _ADJ_COLS)
    if not col_meta:
        raise ValueError("Could not parse period dates from Table 1 header rows")

    pct_index = _index_pct_table(t2_rows)

    # Determine which period corresponds to advance and preliminary for pct lookup.
    # Advance = first col (lowest index, not revised), preliminary = second.
    non_revised = [m for m in col_meta if not m["revised"]]
    adv_col = non_revised[0]["col_idx"] if len(non_revised) >= 1 else None
    pre_col = non_revised[1]["col_idx"] if len(non_revised) >= 2 else None

    records: list[dict] = []
    pending_prefix = ""
    for row in t1_rows[_DATA_START_ROW:]:
        if len(row) <= max(_ADJ_COLS):
            pending_prefix = ""
            continue

        name_raw = row[_NAME_COL]
        has_adj_data = any(
            isinstance(row[c], (int, float)) for c in _ADJ_COLS
        )

        if name_raw and not has_adj_data:
            # Name-only row: accumulate as prefix for the next data row.
            pending_prefix = _clean_name(str(name_raw))
            continue

        if not name_raw and not has_adj_data:
            # Blank separator row: reset prefix.
            pending_prefix = ""
            continue

        # Data row — combine any pending prefix with this row's name fragment.
        name_part = _clean_name(str(name_raw)) if name_raw else ""
        if pending_prefix and name_part:
            name = pending_prefix + " " + name_part
        elif pending_prefix:
            name = pending_prefix
        else:
            name = name_part
        pending_prefix = ""

        name = _DOT_LEADER.sub("", name).strip()
        if not name:
            continue

        pct_data = pct_index.get(name, {})

        for meta in col_meta:
            col_idx = meta["col_idx"]
            sales = _parse_numeric(row[col_idx])
            if sales is None:
                continue

            if col_idx == adv_col:
                mom = pct_data.get("adv_mom")
                yoy = pct_data.get("adv_yoy")
            elif col_idx == pre_col:
                mom = pct_data.get("pre_mom")
                yoy = pct_data.get("pre_yoy")
            else:
                mom = None
                yoy = None

            records.append({
                "series_name": name,
                "period_date": meta["period_date"],
                "sales_millions_usd": sales,
                "month_over_month_pct": mom,
                "year_over_year_pct": yoy,
                "revised": meta["revised"],
                "source_url": source_url,
            })

    if not records:
        raise ValueError("No records extracted from Census retail sales workbook")

    return records


def scrape() -> list[dict]:
    """Fetch the live Census advance retail Excel and return parsed records.

    Delegates HTTP fetching to src.scrapers.http_client.fetch, which enforces
    robots.txt compliance, a polite 2-5 s delay, and exponential backoff on
    429/5xx. An additional 3-second courtesy sleep is applied after receipt.

    Returns:
        Same structure as run().

    Raises:
        ValueError: Propagated from run() when no records can be extracted.
        RuntimeError: When robots.txt disallows the target URL.
    """
    resp = fetch(SOURCE_URL)
    time.sleep(3)
    return run(resp.content, source_url=SOURCE_URL)


def _record_to_proto(record: dict) -> CensusRetailSalesRecord:
    """Convert a parsed record dict to a CensusRetailSalesRecord instance.

    None values for percent-change fields become 0.0 (proto3 default).

    Args:
        record: Dict as returned by run().

    Returns:
        Populated CensusRetailSalesRecord with fetch_time set to current UTC.
    """
    msg = CensusRetailSalesRecord()
    msg.series_name = record["series_name"]
    msg.period_date = record["period_date"]
    msg.sales_millions_usd = record["sales_millions_usd"]
    msg.month_over_month_pct = record["month_over_month_pct"] or 0.0
    msg.year_over_year_pct = record["year_over_year_pct"] or 0.0
    msg.revised = record["revised"]
    msg.source_url = record["source_url"]
    msg.fetch_time = datetime.now(timezone.utc).isoformat()
    return msg


def main() -> int:
    """Scrape Census advance retail sales data and upload records to BigQuery.

    Returns:
        Count of successfully inserted rows.

    Raises:
        ValueError: When scrape() extracts zero records.
    """
    records = scrape()
    messages = [_record_to_proto(r) for r in records]
    return upload_rows("census_retail_sales", messages, date_column="period_date")


if __name__ == "__main__":
    main()
