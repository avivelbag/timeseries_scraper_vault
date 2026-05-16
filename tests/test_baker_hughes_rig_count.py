import os
import re
import sys
import threading
from datetime import datetime
from http.server import BaseHTTPRequestHandler, HTTPServer
from unittest.mock import MagicMock, patch

import openpyxl
import pytest
import requests

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from src.scrapers.baker_hughes_rig_count import (
    REQUIRED_FIELDS,
    SOURCE_URL,
    _find_excel_url,
    _normalise_drill_type,
    _normalise_region,
    _parse_date,
    _safe_int,
    parse_workbook,
    scrape,
)
from protos.baker_hughes_rig_count_pb2 import RigCountRecord
from src.scrapers.http_client import _robots

FIXTURE_PATH = os.path.join(
    os.path.dirname(__file__), "fixtures", "baker_hughes_rig_count_sample.xlsx"
)

_DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _make_workbook(
    sheet_title: str = "North America Rotary Rig Count (Baker Hughes)",
    headers: list | None = None,
    rows: list | None = None,
) -> openpyxl.Workbook:
    if headers is None:
        headers = ["PublishDate", "LandSea", "Location", "DrillFor", "Count", "PriorWeek", "YearAgo"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws.append(headers)
    for row in (rows or []):
        ws.append(row)
    return wb


def _fixture_bytes() -> bytes:
    with open(FIXTURE_PATH, "rb") as fh:
        return fh.read()


def _mock_excel_response(raw_bytes: bytes) -> MagicMock:
    resp = MagicMock(spec=requests.Response)
    resp.iter_content = lambda chunk_size=65536: iter([raw_bytes])
    return resp


# ---------------------------------------------------------------------------
# Tests: _parse_date
# ---------------------------------------------------------------------------


class TestParseDate:
    def test_datetime_object(self):
        assert _parse_date(datetime(2024, 1, 5)) == "2024-01-05"

    def test_iso_string(self):
        assert _parse_date("2024-01-05") == "2024-01-05"

    def test_us_slash_format(self):
        assert _parse_date("01/05/2024") == "2024-01-05"

    def test_two_digit_year(self):
        assert _parse_date("01/05/24") == "2024-01-05"

    def test_none_on_garbage(self):
        assert _parse_date("not-a-date") is None

    def test_none_on_none(self):
        assert _parse_date(None) is None

    def test_none_on_integer(self):
        assert _parse_date(42) is None


# ---------------------------------------------------------------------------
# Tests: _normalise_region
# ---------------------------------------------------------------------------


class TestNormaliseRegion:
    def test_us_exact(self):
        assert _normalise_region("US") == "us"

    def test_us_lowercase(self):
        assert _normalise_region("us") == "us"

    def test_can_upper(self):
        assert _normalise_region("CAN") == "canada"

    def test_canada_full(self):
        assert _normalise_region("canada") == "canada"

    def test_unknown_returns_none(self):
        assert _normalise_region("Mexico") is None

    def test_whitespace_stripped(self):
        assert _normalise_region("  US  ") == "us"


# ---------------------------------------------------------------------------
# Tests: _normalise_drill_type
# ---------------------------------------------------------------------------


class TestNormaliseDrillType:
    def test_oil(self):
        assert _normalise_drill_type("Oil") == "oil"

    def test_gas(self):
        assert _normalise_drill_type("Gas") == "gas"

    def test_misc(self):
        assert _normalise_drill_type("Misc") == "misc"

    def test_miscellaneous(self):
        assert _normalise_drill_type("Miscellaneous") == "misc"

    def test_unknown_returns_none(self):
        assert _normalise_drill_type("Steam") is None


# ---------------------------------------------------------------------------
# Tests: _safe_int
# ---------------------------------------------------------------------------


class TestSafeInt:
    def test_integer(self):
        assert _safe_int(480) == 480

    def test_float(self):
        assert _safe_int(480.0) == 480

    def test_string_numeric(self):
        assert _safe_int("480") == 480

    def test_none_returns_none(self):
        assert _safe_int(None) is None

    def test_non_numeric_returns_none(self):
        assert _safe_int("N/A") is None


# ---------------------------------------------------------------------------
# Tests: _find_excel_url
# ---------------------------------------------------------------------------


class TestFindExcelUrl:
    def test_finds_absolute_xlsx_link(self):
        html = '<html><body><a href="https://example.com/data.xlsx">Download</a></body></html>'
        assert _find_excel_url(html, SOURCE_URL) == "https://example.com/data.xlsx"

    def test_finds_relative_xlsx_link(self):
        html = '<html><body><a href="/files/rig_count.xlsx">Download</a></body></html>'
        result = _find_excel_url(html, "https://rigcount.bakerhughes.com/na-rig-count")
        assert result == "https://rigcount.bakerhughes.com/files/rig_count.xlsx"

    def test_xls_extension_also_matched(self):
        html = '<html><body><a href="https://example.com/data.xls">Download</a></body></html>'
        assert _find_excel_url(html, SOURCE_URL) == "https://example.com/data.xls"

    def test_raises_when_no_link(self):
        html = '<html><body><p>No links here.</p></body></html>'
        with pytest.raises(ValueError, match="No Excel download link found"):
            _find_excel_url(html, SOURCE_URL)

    def test_ignores_non_excel_links(self):
        html = (
            '<html><body>'
            '<a href="https://example.com/report.pdf">PDF</a>'
            '<a href="https://example.com/data.xlsx">Excel</a>'
            '</body></html>'
        )
        assert _find_excel_url(html, SOURCE_URL) == "https://example.com/data.xlsx"


# ---------------------------------------------------------------------------
# Tests: parse_workbook — happy path with fixture
# ---------------------------------------------------------------------------


class TestParseWorkbookFixture:
    @pytest.fixture(scope="class")
    def fixture_records(self):
        wb = openpyxl.load_workbook(FIXTURE_PATH, read_only=True, data_only=True)
        records = parse_workbook(wb, SOURCE_URL)
        wb.close()
        return records

    def test_correct_row_count(self, fixture_records):
        assert len(fixture_records) == 12

    def test_all_required_string_fields_populated(self, fixture_records):
        string_fields = [f for f in REQUIRED_FIELDS if f not in ("rig_count", "week_over_week_change", "year_ago_count")]
        for rec in fixture_records:
            for field in string_fields:
                val = getattr(rec, field)
                assert val not in (None, ""), f"Field '{field}' empty in {rec}"

    def test_report_date_format(self, fixture_records):
        for rec in fixture_records:
            assert _DATE_RE.match(rec.report_date), f"Bad date: {rec.report_date}"

    def test_regions_are_canonical(self, fixture_records):
        regions = {rec.region for rec in fixture_records}
        assert regions == {"us", "canada"}

    def test_drill_types_are_canonical(self, fixture_records):
        drill_types = {rec.drill_type for rec in fixture_records}
        assert drill_types == {"oil", "gas", "misc"}

    def test_rig_counts_are_positive(self, fixture_records):
        for rec in fixture_records:
            assert rec.rig_count > 0, f"Non-positive rig_count in {rec}"

    def test_week_over_week_change_computed(self, fixture_records):
        us_oil = [r for r in fixture_records if r.report_date == "2024-01-05" and r.region == "us" and r.drill_type == "oil"]
        assert len(us_oil) == 1
        assert us_oil[0].week_over_week_change == 2

    def test_source_url_in_every_record(self, fixture_records):
        for rec in fixture_records:
            assert rec.source_url == SOURCE_URL

    def test_fetch_time_is_iso8601(self, fixture_records):
        for rec in fixture_records:
            datetime.fromisoformat(rec.fetch_time)

    def test_all_records_are_rig_count_record_instances(self, fixture_records):
        for rec in fixture_records:
            assert isinstance(rec, RigCountRecord)


# ---------------------------------------------------------------------------
# Tests: parse_workbook — edge cases
# ---------------------------------------------------------------------------


class TestParseWorkbookEdgeCases:
    def test_empty_data_rows_returns_empty_list(self):
        wb = _make_workbook(rows=[])
        records = parse_workbook(wb, SOURCE_URL)
        assert records == []

    def test_unknown_region_row_skipped(self):
        wb = _make_workbook(rows=[
            [datetime(2024, 1, 5), "Land", "Mexico", "Oil", 10, 9, 8],
            [datetime(2024, 1, 5), "Land", "US",     "Oil", 480, 478, 510],
        ])
        records = parse_workbook(wb, SOURCE_URL)
        assert len(records) == 1
        assert records[0].region == "us"

    def test_unknown_drill_type_row_skipped(self):
        wb = _make_workbook(rows=[
            [datetime(2024, 1, 5), "Land", "US", "Steam", 10, 9, 8],
            [datetime(2024, 1, 5), "Land", "US", "Oil",  480, 478, 510],
        ])
        records = parse_workbook(wb, SOURCE_URL)
        assert len(records) == 1
        assert records[0].drill_type == "oil"

    def test_null_count_row_skipped(self):
        wb = _make_workbook(rows=[
            [datetime(2024, 1, 5), "Land", "US", "Oil", None, 478, 510],
            [datetime(2024, 1, 5), "Land", "US", "Gas", 118,  119, 125],
        ])
        records = parse_workbook(wb, SOURCE_URL)
        assert len(records) == 1
        assert records[0].drill_type == "gas"

    def test_unparseable_date_row_skipped(self):
        wb = _make_workbook(rows=[
            ["not-a-date", "Land", "US", "Oil", 480, 478, 510],
            [datetime(2024, 1, 5), "Land", "US", "Gas", 118, 119, 125],
        ])
        records = parse_workbook(wb, SOURCE_URL)
        assert len(records) == 1

    def test_missing_priorweek_column_defaults_to_zero(self):
        wb = _make_workbook(
            headers=["PublishDate", "LandSea", "Location", "DrillFor", "Count", "YearAgo"],
            rows=[[datetime(2024, 1, 5), "Land", "US", "Oil", 480, 510]],
        )
        records = parse_workbook(wb, SOURCE_URL)
        assert len(records) == 1
        assert records[0].week_over_week_change == 480

    def test_missing_yearago_column_defaults_to_zero(self):
        wb = _make_workbook(
            headers=["PublishDate", "LandSea", "Location", "DrillFor", "Count", "PriorWeek"],
            rows=[[datetime(2024, 1, 5), "Land", "US", "Oil", 480, 478]],
        )
        records = parse_workbook(wb, SOURCE_URL)
        assert len(records) == 1
        assert records[0].year_ago_count == 0

    def test_no_north_america_sheet_raises(self):
        wb = openpyxl.Workbook()
        wb.active.title = "Summary"
        with pytest.raises(ValueError, match="No 'North America' sheet found"):
            parse_workbook(wb, SOURCE_URL)

    def test_missing_required_header_columns_raises(self):
        wb = _make_workbook(headers=["Date", "Region", "Type", "Rigs"])
        with pytest.raises(ValueError, match="Required columns not found"):
            parse_workbook(wb, SOURCE_URL)

    def test_large_workbook(self):
        rows = []
        for i in range(100):
            date = datetime(2024, 1, 1 + (i % 28))
            region = "US" if i % 2 == 0 else "CAN"
            drill = ["Oil", "Gas", "Misc"][i % 3]
            count = 100 + i
            prior = 99 + i
            year_ago = 95 + i
            rows.append([date, "Land", region, drill, count, prior, year_ago])
        wb = _make_workbook(rows=rows)
        records = parse_workbook(wb, SOURCE_URL)
        assert len(records) == 100
        for rec in records:
            assert rec.rig_count > 0
            assert _DATE_RE.match(rec.report_date)

    def test_none_location_row_skipped(self):
        wb = _make_workbook(rows=[
            [datetime(2024, 1, 5), "Land", None, "Oil", 480, 478, 510],
            [datetime(2024, 1, 5), "Land", "US", "Oil", 482, 480, 512],
        ])
        records = parse_workbook(wb, SOURCE_URL)
        assert len(records) == 1

    def test_null_priorweek_value_treated_as_zero(self):
        wb = _make_workbook(rows=[
            [datetime(2024, 1, 5), "Land", "US", "Oil", 480, None, 510],
        ])
        records = parse_workbook(wb, SOURCE_URL)
        assert len(records) == 1
        assert records[0].week_over_week_change == 480


# ---------------------------------------------------------------------------
# Tests: RigCountRecord schema / proto defaults
# ---------------------------------------------------------------------------


class TestRigCountRecordSchema:
    def test_default_instance_has_empty_strings(self):
        rec = RigCountRecord()
        assert rec.report_date == ""
        assert rec.region == ""
        assert rec.drill_type == ""
        assert rec.source_url == ""
        assert rec.fetch_time == ""

    def test_default_instance_has_zero_ints(self):
        rec = RigCountRecord()
        assert rec.rig_count == 0
        assert rec.week_over_week_change == 0
        assert rec.year_ago_count == 0

    def test_all_required_fields_on_populated_record(self):
        rec = RigCountRecord(
            report_date="2024-01-05",
            region="us",
            drill_type="oil",
            rig_count=480,
            week_over_week_change=2,
            year_ago_count=510,
            source_url=SOURCE_URL,
            fetch_time="2024-01-05T12:00:00+00:00",
        )
        for field in REQUIRED_FIELDS:
            val = getattr(rec, field)
            assert val is not None


# ---------------------------------------------------------------------------
# Tests: scrape() — mocked network
# ---------------------------------------------------------------------------


class TestScrapeFunction:
    @pytest.fixture(autouse=True)
    def clear_robot_cache(self):
        _robots.cache_clear()
        yield
        _robots.cache_clear()

    def _make_page_response(self, excel_url: str) -> MagicMock:
        html = f'<html><body><a href="{excel_url}">Download Excel</a></body></html>'
        resp = MagicMock(spec=requests.Response)
        resp.text = html
        return resp

    def _make_excel_response(self) -> MagicMock:
        return _mock_excel_response(_fixture_bytes())

    def test_scrape_makes_two_fetch_calls(self):
        excel_url = "https://rigcount.bakerhughes.com/files/rig_count.xlsx"
        page_resp = self._make_page_response(excel_url)
        excel_resp = self._make_excel_response()

        with patch("src.scrapers.baker_hughes_rig_count.fetch", side_effect=[page_resp, excel_resp]) as mock_fetch:
            scrape()

        assert mock_fetch.call_count == 2
        first_call_args = mock_fetch.call_args_list[0]
        assert first_call_args[0][0] == SOURCE_URL

    def test_scrape_returns_correct_record_count(self):
        excel_url = "https://rigcount.bakerhughes.com/files/rig_count.xlsx"
        page_resp = self._make_page_response(excel_url)
        excel_resp = self._make_excel_response()

        with patch("src.scrapers.baker_hughes_rig_count.fetch", side_effect=[page_resp, excel_resp]):
            records = scrape()

        assert len(records) == 12

    def test_scrape_records_have_excel_source_url(self):
        excel_url = "https://rigcount.bakerhughes.com/files/rig_count.xlsx"
        page_resp = self._make_page_response(excel_url)
        excel_resp = self._make_excel_response()

        with patch("src.scrapers.baker_hughes_rig_count.fetch", side_effect=[page_resp, excel_resp]):
            records = scrape()

        for rec in records:
            assert rec.source_url == excel_url

    def test_scrape_uses_min_delay_3s(self):
        excel_url = "https://rigcount.bakerhughes.com/files/rig_count.xlsx"
        page_resp = self._make_page_response(excel_url)
        excel_resp = self._make_excel_response()

        with patch("src.scrapers.baker_hughes_rig_count.fetch", side_effect=[page_resp, excel_resp]) as mock_fetch:
            scrape()

        for call in mock_fetch.call_args_list:
            kwargs = call[1]
            assert kwargs.get("min_delay", 0) >= 3.0

    def test_scrape_page_raises_on_no_excel_link(self):
        resp = MagicMock(spec=requests.Response)
        resp.text = "<html><body>No links here.</body></html>"
        with patch("src.scrapers.baker_hughes_rig_count.fetch", return_value=resp):
            with pytest.raises(ValueError, match="No Excel download link found"):
                scrape()


# ---------------------------------------------------------------------------
# Tests: robots.txt enforcement
# ---------------------------------------------------------------------------


class TestRobotsTxtEnforcement:
    @pytest.fixture(autouse=True)
    def clear_robot_cache(self):
        _robots.cache_clear()
        yield
        _robots.cache_clear()

    @pytest.fixture()
    def disallow_server(self):
        class Handler(BaseHTTPRequestHandler):
            def do_GET(self):
                body = b"User-agent: *\nDisallow: /\n"
                self.send_response(200)
                self.send_header("Content-Type", "text/plain")
                self.send_header("Content-Length", str(len(body)))
                self.end_headers()
                self.wfile.write(body)

            def log_message(self, *args):
                pass

        server = HTTPServer(("127.0.0.1", 0), Handler)
        port = server.server_address[1]
        thread = threading.Thread(target=server.serve_forever, daemon=True)
        thread.start()
        yield f"http://127.0.0.1:{port}"
        server.shutdown()

    def test_scraper_aborts_on_disallowed_path(self, disallow_server):
        from src.scrapers.http_client import fetch as http_fetch

        test_url = disallow_server + "/na-rig-count"
        with patch("src.scrapers.http_client.time.sleep"):
            with pytest.raises(RuntimeError, match="robots.txt disallows"):
                http_fetch(test_url)

    def test_no_download_attempted_when_blocked(self, disallow_server):
        from src.scrapers.http_client import fetch as http_fetch

        session = MagicMock(spec=requests.Session)
        session.headers = {}
        test_url = disallow_server + "/na-rig-count"
        with patch("src.scrapers.http_client.time.sleep"):
            with pytest.raises(RuntimeError):
                http_fetch(test_url, session=session)
        session.get.assert_not_called()
