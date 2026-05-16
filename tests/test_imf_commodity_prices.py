"""Tests for src/scrapers/imf_commodity_prices.py.

All tests use the local fixture Excel file or inline workbooks — no live
network calls are made.  The robots.txt blocking test uses a local HTTP
server to serve a disallowing robots.txt, per orchestrator guidance that
every scraper must assert it aborts on a disallowed path.
"""

import os
import re
import sys
import threading
from datetime import datetime
from http.server import BaseHTTPRequestHandler, HTTPServer
from unittest.mock import MagicMock, patch

import openpyxl
import pytest
import requests

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from src.scrapers.imf_commodity_prices import (
    SOURCE_URL,
    REQUIRED_FIELDS,
    _parse_date,
    _extract_sheet_meta,
    parse_workbook,
    scrape,
    _record_to_proto,
)
from protos.imf_commodity_price_pb2 import ImfCommodityPrice
from src.scrapers.http_client import _robots

FIXTURE_PATH = os.path.join(
    os.path.dirname(__file__), "fixtures", "imf_commodity_prices_sample.xlsx"
)

_DATE_RE = re.compile(r"^\d{4}-\d{2}$")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _make_inline_workbook(sheets: list[tuple]) -> openpyxl.Workbook:
    """Build an in-memory workbook from a list of (title, name, units, rows).

    Args:
        sheets: Each element is (sheet_title, commodity_name, units, data_rows)
            where data_rows is a list of (date_str YYYY-MM-DD, price) tuples.

    Returns:
        An openpyxl Workbook ready to be passed to parse_workbook.
    """
    wb = openpyxl.Workbook()
    first = True
    for sheet_title, commodity_name, units, data_rows in sheets:
        if first:
            ws = wb.active
            ws.title = sheet_title
            first = False
        else:
            ws = wb.create_sheet(title=sheet_title)
        ws.append([commodity_name, units])
        ws.append(["Date", sheet_title])
        for date_str, price in data_rows:
            ws.append([datetime.strptime(date_str, "%Y-%m-%d"), price])
    return wb


# ---------------------------------------------------------------------------
# Tests: _parse_date
# ---------------------------------------------------------------------------


class TestParseDate:
    def test_datetime_object(self):
        dt = datetime(2024, 3, 1)
        assert _parse_date(dt) == "2024-03"

    def test_iso_string(self):
        assert _parse_date("2024-03-01") == "2024-03"

    def test_datetime_string_with_time(self):
        assert _parse_date("2024-03-01 00:00:00") == "2024-03"

    def test_slash_format(self):
        assert _parse_date("03/01/2024") == "2024-03"

    def test_mon_year_format(self):
        assert _parse_date("Mar-2024") == "2024-03"

    def test_none_on_garbage(self):
        assert _parse_date("not-a-date") is None

    def test_none_on_none(self):
        assert _parse_date(None) is None

    def test_none_on_integer(self):
        assert _parse_date(42) is None


# ---------------------------------------------------------------------------
# Tests: _extract_sheet_meta
# ---------------------------------------------------------------------------


class TestExtractSheetMeta:
    def test_returns_name_code_units(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "POILWTI"
        ws.append(["Crude Oil, WTI", "US Dollars per Barrel"])
        ws.append(["Date", "POILWTI"])
        name, code, units = _extract_sheet_meta(ws)
        assert name == "Crude Oil, WTI"
        assert code == "POILWTI"
        assert units == "US Dollars per Barrel"

    def test_falls_back_to_sheet_title_for_code(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "MYCODE"
        ws.append(["Some Commodity", "USD/unit"])
        ws.append(["Date", None])
        _, code, _ = _extract_sheet_meta(ws)
        assert code == "MYCODE"

    def test_empty_sheet_returns_defaults(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "EMPTY"
        name, code, units = _extract_sheet_meta(ws)
        assert name == ""
        assert code == "EMPTY"
        assert units == ""


# ---------------------------------------------------------------------------
# Tests: parse_workbook — happy path with fixture
# ---------------------------------------------------------------------------


class TestParseWorkbookFixture:
    @pytest.fixture(scope="class")
    def fixture_records(self):
        wb = openpyxl.load_workbook(FIXTURE_PATH, read_only=True, data_only=True)
        return parse_workbook(wb, SOURCE_URL)

    def test_correct_row_count(self, fixture_records):
        """Fixture has 2 sheets × 5 data rows = 10 records."""
        assert len(fixture_records) == 10

    def test_all_required_fields_populated(self, fixture_records):
        for rec in fixture_records:
            for field in REQUIRED_FIELDS:
                assert rec.get(field) not in (None, "", 0), (
                    f"Field '{field}' is missing or empty in record: {rec}"
                )

    def test_prices_are_positive_floats(self, fixture_records):
        for rec in fixture_records:
            assert isinstance(rec["price_usd"], float)
            assert rec["price_usd"] > 0.0

    def test_date_strings_match_yyyy_mm(self, fixture_records):
        for rec in fixture_records:
            assert _DATE_RE.match(rec["date"]), f"Bad date: {rec['date']}"

    def test_source_url_in_every_record(self, fixture_records):
        for rec in fixture_records:
            assert rec["source_url"] == SOURCE_URL

    def test_both_commodity_codes_present(self, fixture_records):
        codes = {rec["commodity_code"] for rec in fixture_records}
        assert "POILWTI" in codes
        assert "PCOALAU" in codes

    def test_fetch_time_is_iso8601(self, fixture_records):
        for rec in fixture_records:
            datetime.fromisoformat(rec["fetch_time"])


# ---------------------------------------------------------------------------
# Tests: parse_workbook — edge cases
# ---------------------------------------------------------------------------


class TestParseWorkbookEdgeCases:
    def test_empty_workbook_returns_empty_list(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "EmptySheet"
        records = parse_workbook(wb, SOURCE_URL)
        assert records == []

    def test_non_numeric_price_row_is_skipped(self):
        wb = _make_inline_workbook(
            [("POILWTI", "Oil", "USD", [("2024-01-01", "N/A"), ("2024-02-01", 75.0)])]
        )
        records = parse_workbook(wb, SOURCE_URL)
        assert len(records) == 1
        assert records[0]["date"] == "2024-02"

    def test_zero_price_row_is_skipped(self):
        wb = _make_inline_workbook(
            [("POILWTI", "Oil", "USD", [("2024-01-01", 0.0), ("2024-02-01", 80.0)])]
        )
        records = parse_workbook(wb, SOURCE_URL)
        assert len(records) == 1
        assert records[0]["price_usd"] == pytest.approx(80.0)

    def test_negative_price_row_is_skipped(self):
        wb = _make_inline_workbook(
            [("POILWTI", "Oil", "USD", [("2024-01-01", -10.0), ("2024-02-01", 70.0)])]
        )
        records = parse_workbook(wb, SOURCE_URL)
        assert len(records) == 1

    def test_skip_sheet_names_are_excluded(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Index"
        ws.append(["Index Sheet", ""])
        ws.append(["Date", "N/A"])
        ws.append([datetime(2024, 1, 1), 100.0])
        records = parse_workbook(wb, SOURCE_URL)
        assert records == []

    def test_unparseable_date_row_skipped(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "PCOALAU"
        ws.append(["Coal", "USD/ton"])
        ws.append(["Date", "PCOALAU"])
        ws.append(["not-a-date", 120.0])
        ws.append([datetime(2024, 3, 1), 130.0])
        records = parse_workbook(wb, SOURCE_URL)
        assert len(records) == 1
        assert records[0]["date"] == "2024-03"

    def test_large_workbook_all_records_valid(self):
        """3 sheets × 12 months = 36 records, all must pass field checks."""
        months = [datetime(2024, m, 1) for m in range(1, 13)]
        sheets = [
            ("POILWTI", "WTI Crude", "USD/bbl", [(d.strftime("%Y-%m-%d"), 75.0 + i) for i, d in enumerate(months)]),
            ("PCOALAU", "Coal", "USD/mt", [(d.strftime("%Y-%m-%d"), 130.0 + i) for i, d in enumerate(months)]),
            ("PNGAJPN", "Natural Gas Japan", "USD/mmbtu", [(d.strftime("%Y-%m-%d"), 12.0 + i) for i, d in enumerate(months)]),
        ]
        wb = _make_inline_workbook(sheets)
        records = parse_workbook(wb, SOURCE_URL)
        assert len(records) == 36
        for rec in records:
            assert rec["price_usd"] > 0
            assert _DATE_RE.match(rec["date"])


# ---------------------------------------------------------------------------
# Tests: proto conversion
# ---------------------------------------------------------------------------


class TestRecordToProto:
    def test_all_fields_mapped(self):
        rec = {
            "commodity_name": "Crude Oil, WTI",
            "commodity_code": "POILWTI",
            "date": "2024-03",
            "price_usd": 81.35,
            "units": "US Dollars per Barrel",
            "source_url": SOURCE_URL,
            "fetch_time": "2024-03-01T00:00:00+00:00",
        }
        msg = _record_to_proto(rec)
        assert isinstance(msg, ImfCommodityPrice)
        assert msg.commodity_name == "Crude Oil, WTI"
        assert msg.commodity_code == "POILWTI"
        assert msg.date == "2024-03"
        assert msg.price_usd == pytest.approx(81.35)
        assert msg.units == "US Dollars per Barrel"
        assert msg.source_url == SOURCE_URL
        assert msg.fetch_time == "2024-03-01T00:00:00+00:00"

    def test_proto_dataclass_defaults(self):
        msg = ImfCommodityPrice()
        assert msg.commodity_name == ""
        assert msg.commodity_code == ""
        assert msg.date == ""
        assert msg.price_usd == 0.0
        assert msg.units == ""
        assert msg.source_url == ""
        assert msg.fetch_time == ""


# ---------------------------------------------------------------------------
# Tests: scrape() — mocked network
# ---------------------------------------------------------------------------


class TestScrapeFunction:
    @pytest.fixture(autouse=True)
    def clear_robot_cache(self):
        _robots.cache_clear()
        yield
        _robots.cache_clear()

    def _make_fixture_response(self):
        """Return a mock requests.Response that streams the fixture Excel bytes."""
        with open(FIXTURE_PATH, "rb") as fh:
            raw_bytes = fh.read()

        resp = MagicMock(spec=requests.Response)
        resp.iter_content = lambda chunk_size=65536: iter([raw_bytes])
        return resp

    def test_scrape_calls_fetch_with_source_url(self):
        fake_resp = self._make_fixture_response()
        with patch("src.scrapers.imf_commodity_prices.fetch", return_value=fake_resp) as mock_fetch:
            records = scrape()
        mock_fetch.assert_called_once_with(SOURCE_URL, stream=True, min_delay=3.0, max_delay=6.0)
        assert len(records) == 10

    def test_scrape_no_live_network(self):
        fake_resp = self._make_fixture_response()
        with patch("src.scrapers.imf_commodity_prices.fetch", return_value=fake_resp) as mock_fetch:
            scrape()
        assert mock_fetch.call_count == 1

    def test_scrape_returns_all_required_fields(self):
        fake_resp = self._make_fixture_response()
        with patch("src.scrapers.imf_commodity_prices.fetch", return_value=fake_resp):
            records = scrape()
        for rec in records:
            for field in REQUIRED_FIELDS:
                assert rec.get(field) not in (None, ""), f"Missing {field}"


# ---------------------------------------------------------------------------
# Tests: robots.txt enforcement
# ---------------------------------------------------------------------------


class TestRobotsTxtEnforcement:
    """Assert the scraper aborts when robots.txt disallows the path.

    Uses a real in-process HTTP server that serves a disallowing robots.txt,
    so no live network traffic leaves the machine.
    """

    @pytest.fixture(autouse=True)
    def clear_robot_cache(self):
        _robots.cache_clear()
        yield
        _robots.cache_clear()

    @pytest.fixture()
    def disallow_server(self):
        """Start a local HTTP server that disallows all paths via robots.txt."""

        class Handler(BaseHTTPRequestHandler):
            def do_GET(self):
                body = b"User-agent: *\nDisallow: /\n"
                self.send_response(200)
                self.send_header("Content-Type", "text/plain")
                self.send_header("Content-Length", str(len(body)))
                self.end_headers()
                self.wfile.write(body)

            def log_message(self, *args):
                pass  # silence server log noise during tests

        server = HTTPServer(("127.0.0.1", 0), Handler)
        port = server.server_address[1]
        thread = threading.Thread(target=server.serve_forever, daemon=True)
        thread.start()
        yield f"http://127.0.0.1:{port}"
        server.shutdown()

    def test_scraper_aborts_on_disallowed_path(self, disallow_server):
        """http_client.fetch must raise RuntimeError when robots.txt forbids the URL."""
        from src.scrapers.http_client import fetch as http_fetch

        test_url = disallow_server + "/data/prices.xlsx"
        with patch("src.scrapers.http_client.time.sleep"):
            with pytest.raises(RuntimeError, match="robots.txt disallows"):
                http_fetch(test_url)

    def test_no_download_attempted_when_blocked(self, disallow_server):
        """Session.get must never be called if robots.txt disallows the path."""
        from src.scrapers.http_client import fetch as http_fetch

        session = MagicMock(spec=requests.Session)
        session.headers = {}
        test_url = disallow_server + "/data/prices.xlsx"
        with patch("src.scrapers.http_client.time.sleep"):
            with pytest.raises(RuntimeError):
                http_fetch(test_url, session=session)
        session.get.assert_not_called()
