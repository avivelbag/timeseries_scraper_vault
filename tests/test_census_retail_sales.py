"""Tests for src/scrapers/census_retail_sales.py.

All tests use the static Excel fixture at tests/fixtures/census_retail_sales.xlsx
or small inline Excel bytes built via openpyxl — zero live network calls.
The fixture is a real copy of the Census marts_current.xlsx advance release.
"""

import io
import os
import sys
from datetime import date
from unittest.mock import MagicMock, patch

import openpyxl
import pytest
import requests

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from src.scrapers.census_retail_sales import (
    SOURCE_URL,
    _clean_name,
    _parse_col_meta,
    _parse_numeric,
    _record_to_proto,
    run,
    scrape,
)
from protos.census_retail_sales_pb2 import CensusRetailSalesRecord

FIXTURE_PATH = os.path.join(os.path.dirname(__file__), "fixtures", "census_retail_sales.xlsx")


@pytest.fixture
def fixture_bytes() -> bytes:
    with open(FIXTURE_PATH, "rb") as fh:
        return fh.read()


# ---------------------------------------------------------------------------
# Happy-path tests against the real fixture
# ---------------------------------------------------------------------------


class TestRunHappyPath:
    def test_returns_at_least_twenty_records(self, fixture_bytes):
        """Fixture must yield the minimum required record count."""
        assert len(run(fixture_bytes)) >= 20

    def test_period_dates_are_iso8601_first_of_month(self, fixture_bytes):
        """Every period_date must be YYYY-MM-01 format."""
        import re
        pattern = re.compile(r"^\d{4}-\d{2}-01$")
        for r in run(fixture_bytes):
            assert pattern.match(r["period_date"]), f"Bad period_date: {r['period_date']!r}"
            parsed = date.fromisoformat(r["period_date"])
            assert parsed.year >= 2020

    def test_sales_millions_usd_is_positive_float(self, fixture_bytes):
        """All emitted sales values must be positive floats."""
        for r in run(fixture_bytes):
            assert isinstance(r["sales_millions_usd"], float)
            assert r["sales_millions_usd"] > 0

    def test_revised_field_is_bool(self, fixture_bytes):
        """revised must be a Python bool, not an int or string."""
        for r in run(fixture_bytes):
            assert isinstance(r["revised"], bool)

    def test_at_least_one_revised_record(self, fixture_bytes):
        """Fixture has (r)-labelled columns, so at least one record must be revised=True."""
        assert any(r["revised"] for r in run(fixture_bytes))

    def test_at_least_one_non_revised_record(self, fixture_bytes):
        """Advance and preliminary periods are not revised."""
        assert any(not r["revised"] for r in run(fixture_bytes))

    def test_advance_period_has_mom_and_yoy(self, fixture_bytes):
        """The advance estimate should have MoM and YoY percent changes from Table 2."""
        records = run(fixture_bytes)
        non_revised = [r for r in records if not r["revised"]]
        with_mom = [r for r in non_revised if r["month_over_month_pct"] is not None]
        assert len(with_mom) > 0, "No non-revised record has MoM pct from Table 2"

    def test_known_total_advance_apr_2026(self, fixture_bytes):
        """Retail & food services total advance Apr 2026 must match the press release."""
        records = run(fixture_bytes)
        match = [
            r for r in records
            if "Retail & food services, total" in r["series_name"]
            and r["period_date"] == "2026-04-01"
            and not r["revised"]
        ]
        assert len(match) == 1
        assert abs(match[0]["sales_millions_usd"] - 757085.0) < 1.0
        assert abs(match[0]["month_over_month_pct"] - 0.5) < 0.01
        assert abs(match[0]["year_over_year_pct"] - 4.9) < 0.01

    def test_revised_period_has_no_pct_changes(self, fixture_bytes):
        """Revised-period records (Feb 2026 and earlier) have no pct data in Table 2."""
        records = run(fixture_bytes)
        feb_revised = [
            r for r in records
            if r["revised"] and r["period_date"] == "2026-02-01"
        ]
        assert len(feb_revised) > 0
        assert all(r["month_over_month_pct"] is None for r in feb_revised)

    def test_source_url_stored_in_every_record(self, fixture_bytes):
        for r in run(fixture_bytes, source_url="https://test.example/"):
            assert r["source_url"] == "https://test.example/"

    def test_series_names_have_no_dot_leaders(self, fixture_bytes):
        """Census dot-leaders (…) and trailing dots must be stripped from series names."""
        for r in run(fixture_bytes):
            assert "…" not in r["series_name"]
            assert not r["series_name"].rstrip().endswith(".")

    def test_multiple_series_present(self, fixture_bytes):
        """Fixture should have more than one unique series (NAICS category)."""
        series = {r["series_name"] for r in run(fixture_bytes)}
        assert len(series) >= 10

    def test_multiple_period_dates_present(self, fixture_bytes):
        """Fixture covers advance, preliminary, and revised periods."""
        dates = {r["period_date"] for r in run(fixture_bytes)}
        assert len(dates) >= 3


# ---------------------------------------------------------------------------
# Edge-case tests
# ---------------------------------------------------------------------------


class TestRunEdgeCases:
    def test_empty_bytes_raises_value_error(self):
        """run() must raise ValueError when given empty bytes."""
        with pytest.raises(ValueError, match="[Ee]mpty"):
            run(b"")

    def test_wrong_sheet_names_raises_value_error(self):
        """run() must raise ValueError when the workbook lacks Table 1. or Table 2."""
        wb = openpyxl.Workbook()
        wb.active.title = "WrongSheet"
        buf = io.BytesIO()
        wb.save(buf)
        with pytest.raises(ValueError, match="Table 1"):
            run(buf.getvalue())

    def test_suppressed_cells_return_none(self):
        """_parse_numeric handles Census suppression codes as None."""
        assert _parse_numeric("(*)") is None
        assert _parse_numeric("(NA)") is None
        assert _parse_numeric("(S)") is None
        assert _parse_numeric(None) is None
        assert _parse_numeric("") is None
        assert _parse_numeric(" ") is None

    def test_dash_cells_return_none(self):
        assert _parse_numeric("-") is None
        assert _parse_numeric("--") is None

    def test_numeric_cells_parse_correctly(self):
        assert _parse_numeric(757085) == 757085.0
        assert _parse_numeric(0.5) == 0.5
        assert _parse_numeric("1,234") == 1234.0
        assert abs(_parse_numeric(-3.0) - (-3.0)) < 0.001

    def test_clean_name_strips_dot_leaders(self):
        """_clean_name removes Census dot-leader suffixes from series labels."""
        assert _clean_name("Retail & food services……..….") == "Retail & food services"
        assert _clean_name("  Motor vehicle & parts dealers …….………..") == "Motor vehicle & parts dealers"
        assert _clean_name(None) == ""
        assert _clean_name("") == ""

    def test_multiline_series_name_concatenation(self, fixture_bytes):
        """Series spanning two rows (e.g., 'Total (excl. motor vehicle & parts &\\ngasoline stations)')
        must appear as a single assembled series_name."""
        records = run(fixture_bytes)
        names = {r["series_name"] for r in records}
        assembled = [n for n in names if "gasoline stations" in n.lower() and "excl" in n.lower()]
        assert len(assembled) >= 1, f"Multi-line assembled name not found. Names: {sorted(names)[:10]}"


# ---------------------------------------------------------------------------
# Failure-mode tests
# ---------------------------------------------------------------------------


class TestRunFailureModes:
    def test_no_data_rows_raises_value_error(self):
        """A workbook with empty tables must raise ValueError."""
        wb = openpyxl.Workbook()
        wb.active.title = "Table 1."
        wb.create_sheet("Table 2.")
        buf = io.BytesIO()
        wb.save(buf)
        with pytest.raises(ValueError):
            run(buf.getvalue())

    def test_missing_table2_raises_value_error(self):
        """A workbook with only Table 1 must raise ValueError."""
        wb = openpyxl.Workbook()
        wb.active.title = "Table 1."
        buf = io.BytesIO()
        wb.save(buf)
        with pytest.raises(ValueError):
            run(buf.getvalue())

    def test_all_suppressed_cells_yields_no_records(self):
        """When all adjusted-section cells are suppressed, run raises ValueError."""
        from src.scrapers.census_retail_sales import (
            _YEAR_ROW_IDX, _DATA_START_ROW,
        )
        wb = openpyxl.Workbook()
        t1 = wb.active
        t1.title = "Table 1."
        t2 = wb.create_sheet("Table 2.")

        # Build Table 1 with header band at the correct row indices.
        empty = [None] * 14
        while t1.max_row < _YEAR_ROW_IDX:
            t1.append(empty)
        # Row at _YEAR_ROW_IDX (0-based) = Excel row _YEAR_ROW_IDX+1
        t1.append(["NAICS", "Kind of Business", None, None, 2026, None, None, 2025, None, 2026, None, None, 2025, None])
        t1.append(["code", None, None, "% Chg.", "Apr.", "Mar.", "Feb.", "Apr.", "Mar.", "Apr.", "Mar.", "Feb.", "Apr.", "Mar."])
        t1.append([None, None, 2026, 2025, "(a)", "(p)", "(r)", None, None, "(a)", "(p)", "(r)", "(r)", "(r)"])
        # Pad rows up to _DATA_START_ROW
        while t1.max_row < _DATA_START_ROW:
            t1.append(empty)
        # A data row where all adjusted cols (9-13) are suppressed
        t1.append([441, "Motor vehicle & parts dealers", None, None, "(*)", "(*)", "(*)", None, None, "(*)", "(*)", "(*)", "(*)", "(*)"])

        t2.append([None, "Kind of Business", "Adv MoM", "Adv YoY", "Pre MoM", "Pre YoY"])
        for _ in range(8):
            t2.append([None] * 6)
        t2.append([441, "Motor vehicle & parts dealers", None, None, None, None])

        buf = io.BytesIO()
        wb.save(buf)
        with pytest.raises(ValueError, match="No records extracted"):
            run(buf.getvalue())


# ---------------------------------------------------------------------------
# _parse_col_meta unit tests
# ---------------------------------------------------------------------------


class TestParseColMeta:
    def _make_rows(self, year_vals, month_vals, marker_vals, col_indices):
        """Build minimal row list for _parse_col_meta testing."""
        from src.scrapers.census_retail_sales import _YEAR_ROW_IDX, _MONTH_ROW_IDX, _MARKER_ROW_IDX
        rows = [None] * (max(_YEAR_ROW_IDX, _MONTH_ROW_IDX, _MARKER_ROW_IDX) + 1)
        rows[_YEAR_ROW_IDX] = tuple(year_vals)
        rows[_MONTH_ROW_IDX] = tuple(month_vals)
        rows[_MARKER_ROW_IDX] = tuple(marker_vals)
        return rows

    def test_advance_col_is_not_revised(self):
        year_vals = [None] * 14
        year_vals[9] = 2026
        month_vals = [None] * 14
        month_vals[9] = "Apr."
        marker_vals = [None] * 14
        marker_vals[9] = "(a)"
        rows = self._make_rows(year_vals, month_vals, marker_vals, [9])
        meta = _parse_col_meta(rows, [9])
        assert len(meta) == 1
        assert meta[0]["period_date"] == "2026-04-01"
        assert meta[0]["revised"] is False

    def test_revised_col_is_marked(self):
        year_vals = [None] * 14
        year_vals[9] = 2026
        year_vals[11] = None  # carries from 9
        month_vals = [None] * 14
        month_vals[9] = "Feb."
        month_vals[11] = "Feb."
        marker_vals = [None] * 14
        marker_vals[9] = "(a)"
        marker_vals[11] = "(r)"
        rows = self._make_rows(year_vals, month_vals, marker_vals, [9, 11])
        meta = _parse_col_meta(rows, [9, 11])
        results = {m["col_idx"]: m for m in meta}
        assert results[9]["revised"] is False
        assert results[11]["revised"] is True

    def test_year_carries_forward(self):
        """Year value carries forward from a non-None column to subsequent None columns."""
        year_vals = [None] * 14
        year_vals[9] = 2026
        month_vals = [None] * 14
        month_vals[9] = "Apr."
        month_vals[10] = "Mar."
        month_vals[11] = "Feb."
        marker_vals = [None] * 14
        marker_vals[9] = "(a)"
        marker_vals[10] = "(p)"
        marker_vals[11] = "(r)"
        rows = self._make_rows(year_vals, month_vals, marker_vals, [9, 10, 11])
        meta = _parse_col_meta(rows, [9, 10, 11])
        assert len(meta) == 3
        assert all(m["period_date"].startswith("2026-") for m in meta)

    def test_unrecognised_month_skipped(self):
        year_vals = [None] * 14
        year_vals[9] = 2026
        month_vals = [None] * 14
        month_vals[9] = "Qxx"  # invalid
        marker_vals = [None] * 14
        marker_vals[9] = "(a)"
        rows = self._make_rows(year_vals, month_vals, marker_vals, [9])
        meta = _parse_col_meta(rows, [9])
        assert meta == []


# ---------------------------------------------------------------------------
# Proto stub tests
# ---------------------------------------------------------------------------


class TestRecordToProto:
    def _sample_record(self, **overrides) -> dict:
        base = {
            "series_name": "Motor vehicle & parts dealers",
            "period_date": "2026-04-01",
            "sales_millions_usd": 139229.0,
            "month_over_month_pct": -0.4,
            "year_over_year_pct": -1.2,
            "revised": False,
            "source_url": SOURCE_URL,
        }
        base.update(overrides)
        return base

    def test_proto_fields_populated(self, fixture_bytes):
        records = run(fixture_bytes)
        assert records, "Need at least one record from fixture"
        msg = _record_to_proto(records[0])
        assert isinstance(msg, CensusRetailSalesRecord)
        assert msg.series_name == records[0]["series_name"]
        assert msg.period_date == records[0]["period_date"]
        assert isinstance(msg.sales_millions_usd, float)
        assert isinstance(msg.revised, bool)
        assert msg.source_url == records[0]["source_url"]

    def test_none_pct_becomes_zero(self):
        rec = self._sample_record(month_over_month_pct=None, year_over_year_pct=None)
        msg = _record_to_proto(rec)
        assert msg.month_over_month_pct == 0.0
        assert msg.year_over_year_pct == 0.0

    def test_fetch_time_is_iso8601(self, fixture_bytes):
        records = run(fixture_bytes)
        msg = _record_to_proto(records[0])
        assert "T" in msg.fetch_time
        assert msg.fetch_time != ""

    def test_revised_flag_round_trips(self, fixture_bytes):
        records = run(fixture_bytes)
        revised_rec = next(r for r in records if r["revised"])
        non_revised_rec = next(r for r in records if not r["revised"])
        assert _record_to_proto(revised_rec).revised is True
        assert _record_to_proto(non_revised_rec).revised is False

    def test_source_url_non_empty(self, fixture_bytes):
        records = run(fixture_bytes)
        msg = _record_to_proto(records[0])
        assert msg.source_url != ""


# ---------------------------------------------------------------------------
# scrape() integration tests (no live network)
# ---------------------------------------------------------------------------


class TestScrapeFunction:
    def test_scrape_calls_fetch_with_source_url(self, fixture_bytes):
        fake_resp = MagicMock(spec=requests.Response)
        fake_resp.content = fixture_bytes

        with (
            patch("src.scrapers.census_retail_sales.fetch", return_value=fake_resp) as mock_fetch,
            patch("src.scrapers.census_retail_sales.time.sleep"),
        ):
            records = scrape()

        mock_fetch.assert_called_once_with(SOURCE_URL)
        assert len(records) > 0

    def test_scrape_sleeps_at_least_3_seconds(self, fixture_bytes):
        fake_resp = MagicMock(spec=requests.Response)
        fake_resp.content = fixture_bytes
        sleep_calls: list[float] = []

        with (
            patch("src.scrapers.census_retail_sales.fetch", return_value=fake_resp),
            patch(
                "src.scrapers.census_retail_sales.time.sleep",
                side_effect=lambda s: sleep_calls.append(s),
            ),
        ):
            scrape()

        assert any(s >= 3 for s in sleep_calls), f"No sleep >= 3s; calls: {sleep_calls}"

    def test_scrape_returns_same_as_run(self, fixture_bytes):
        fake_resp = MagicMock(spec=requests.Response)
        fake_resp.content = fixture_bytes

        with (
            patch("src.scrapers.census_retail_sales.fetch", return_value=fake_resp),
            patch("src.scrapers.census_retail_sales.time.sleep"),
        ):
            scraped = scrape()

        direct = run(fixture_bytes)
        assert scraped == direct

    def test_scrape_propagates_value_error_on_empty_response(self):
        fake_resp = MagicMock(spec=requests.Response)
        fake_resp.content = b""

        with (
            patch("src.scrapers.census_retail_sales.fetch", return_value=fake_resp),
            patch("src.scrapers.census_retail_sales.time.sleep"),
        ):
            with pytest.raises(ValueError):
                scrape()
